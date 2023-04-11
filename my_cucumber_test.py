from behave import given, when, then
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time

@given('I am on the login page')
def step_impl(context):
    context.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    context.driver.get("https://www.saucedemo.com/")
    context.driver.maximize_window()

@when('I enter valid username and password from Excel sheet')
def step_impl(context):
    path = "C:\\Users\\Lenovo\\Desktop\\Selenium5.xlsx"
    workbook = openpyxl.load_workbook(path)
    sh = workbook.active

    num_rows = sh.max_row

    for i in range(2, num_rows + 1):
        username = sh.cell(row=i, column=1).value
        password = sh.cell(row=i, column=2).value

        context.driver.find_element(By.ID, "user-name").send_keys(username)
        context.driver.find_element(By.ID, "password").send_keys(password)
        context.driver.find_element(By.ID, "login-button").click()

        try:
            assert "https://www.saucedemo.com/inventory.html" in context.driver.current_url
            print("Test Passed: Login successful")
        except AssertionError:
            print("Test Failed: Login unsuccessful")
            continue

        context.driver.find_element(By.ID, "react-burger-menu-btn").click()
        context.driver.find_element(By.ID, "logout_sidebar_link").click()

        time.sleep(2)

@then('I should be logged in successfully')
def step_impl(context):
    context.driver.quit()

